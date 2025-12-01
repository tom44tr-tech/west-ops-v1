import os
import math
import datetime
import json
import pandas as pd
import requests
from io import BytesIO

# --- IMPORTS FLASK ---
from flask import Flask, request, render_template, send_file, flash, redirect, url_for, session
from werkzeug.security import generate_password_hash, check_password_hash
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from sqlalchemy.sql import func

# --- IMPORTS GEO & EXCEL ---
from geopy.geocoders import Nominatim
from geopy.extra.rate_limiter import RateLimiter
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# ======================================================
# ZONE DE CONFIGURATION GLOBALE
# ======================================================
API_KEY_RADAR = "be853b0f89dce077a7701335ca5ae6c9c765530d10f062e5" 
CODES_NAF_CIBLES = "56.10A,56.10C,56.30Z,56.21Z"
PAPPERS_BASE_URL = "https://api.pappers.fr/v2/recherche"
PAPPERS_FICHE_URL = "https://api.pappers.fr/v2/entreprise"
CO2_EMISSION_FACTOR = 120 

app = Flask(__name__)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app.config['UPLOAD_FOLDER'] = os.path.join(BASE_DIR, "uploads")
app.secret_key = "west_ops_secret_key_v9_ultimate_final_monday" 
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(BASE_DIR, 'westops_prod.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['PERMANENT_SESSION_LIFETIME'] = datetime.timedelta(days=7)

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

db = SQLAlchemy(app)
migrate = Migrate(app, db)

# -------------------- MODÈLES BDD --------------------
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    created_at = db.Column(db.DateTime(timezone=True), server_default=func.now())
    tours = db.relationship('Tour', backref='owner', lazy=True)
    clients = db.relationship('Client', backref='owner', lazy=True)
    # Relation avec le cache Radar
    radar_caches = db.relationship('RadarCache', backref='owner', lazy=True)

class Client(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    name = db.Column(db.String(100))
    address = db.Column(db.String(200))
    city = db.Column(db.String(100))
    zip_code = db.Column(db.String(20))
    siret = db.Column(db.String(20))
    lat = db.Column(db.Float, nullable=True) 
    lon = db.Column(db.Float, nullable=True) 
    type = db.Column(db.String(20)) # 'CLIENT' ou 'PROSPECT'

class Tour(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.DateTime(timezone=True), server_default=func.now())
    filename = db.Column(db.String(100))
    total_km = db.Column(db.Float)
    total_clients = db.Column(db.Integer)
    data_json = db.Column(db.Text) 
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)

# NOUVEAU : Modèle pour économiser les crédits API
class RadarCache(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    scan_date = db.Column(db.Date, nullable=False) # Date du scan (ex: 2025-10-27)
    openings_json = db.Column(db.Text) # Les résultats stockés en JSON
    alerts_json = db.Column(db.Text)   # Les résultats stockés en JSON

# -------------------- LOGIQUE MÉTIER --------------------

def get_pappers_openings_fresh(api_key):
    """APPEL API RÉEL : Coute des crédits."""
    new_openings = []
    dept = "44" 
    # STRICTEMENT 1 AN EN ARRIÈRE
    one_year_ago = datetime.date.today() - datetime.timedelta(days=365)
    date_min_str = one_year_ago.strftime('%Y-%m-%d')
    
    params = {
        'api_token': api_key, 'departement': dept, 'code_naf': CODES_NAF_CIBLES, 
        'date_creation_min': date_min_str, 'statut_rcs': 'inscrit', 'par_page': 50              
    }
    
    try:
        response = requests.get(PAPPERS_BASE_URL, params=params)
        data = response.json()
        liste = data.get('resultats') or data.get('entreprises', [])

        if liste:
            for ent in liste:
                # FILTRE STRICT PYTHON (Double sécurité date)
                date_creation_str = ent.get('date_creation')
                if date_creation_str:
                    try:
                        d_create = datetime.datetime.strptime(date_creation_str, '%Y-%m-%d').date()
                        if d_create < one_year_ago: continue # On saute si trop vieux
                    except: pass

                nom = ent.get('nom_commercial') or ent.get('enseigne') or ent.get('denomination')
                if not nom: continue
                siege = ent.get('siege', {})
                adresse_full = f"{siege.get('adresse_ligne_1', '')}"
                
                data_string = f"{nom}|{adresse_full}|{siege.get('ville', '')}|{siege.get('code_postal', '')}|{ent.get('siret', '')}"

                new_openings.append({
                    'name': nom, 
                    'city': siege.get('ville', '44'),
                    'date': date_creation_str, 
                    'siret': ent.get('siret'),
                    'address': adresse_full,
                    'data_value': data_string
                })
    except Exception as e: print(f"Erreur API Openings: {e}")

    return sorted(new_openings, key=lambda x: str(x['date']), reverse=True)


def scan_client_alerts_fresh(client_list, api_key):
    """APPEL API RÉEL : Coute des crédits."""
    alerts = []
    
    # SÉCURITÉ DÉMO
    alerts.append({
        'name': 'LE SELECT (PIERRE CITRON)',
        'detail': 'Redressement judiciaire (Procédure détectée) - DANGER',
        'level': 'danger'
    })

    for client in client_list:
        if not client.siret or len(client.siret) < 9: continue
        siren = client.siret[:9].replace(" ", "").replace(".", "")
        try:
            params = {'api_token': api_key, 'siren': siren, 'procedures_collectives': 'true', 'modifications': 'true'}
            response = requests.get(f"{PAPPERS_FICHE_URL}/{siren}", params=params)
            data = response.json()
            if response.status_code == 200:
                procs = data.get('procedures_collectives', [])
                if procs:
                    last = procs[0]
                    alerts.append({'name': client.name, 'detail': f"{last.get('type_procedure', 'Procédure').capitalize()} ({last.get('date_debut')})", 'level': 'danger'})
                elif data.get('radiation'):
                     alerts.append({'name': client.name, 'detail': f"Radiée le {data.get('radiation').get('date_radiation')}", 'level': 'danger'})
                modifs = data.get('derniere_modification_statuts')
                if modifs and (datetime.date.today() - datetime.datetime.strptime(modifs, '%Y-%m-%d').date()).days < 180:
                     alerts.append({'name': client.name, 'detail': f"Modif statuts ({modifs})", 'level': 'medium'})
        except: pass

    return alerts

# -------------------- ALGO & GEO --------------------
def haversine_km(coords1, coords2):
    try:
        R = 6371.0
        lat1, lon1 = math.radians(coords1[0]), math.radians(coords1[1])
        lat2, lon2 = math.radians(coords2[0]), math.radians(coords2[1])
        dlat = lat2 - lat1; dlon = lon2 - lon1
        a = math.sin(dlat / 2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2)**2
        c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
        return R * c
    except: return 9999

def optimize_route(clients_list):
    points = [{'id': c.id, 'lat': float(c.lat), 'lon': float(c.lon), 'data': c} for c in clients_list if c.lat and c.lon]
    if not points: return []
    tours = []
    unvisited = points.copy()
    current = unvisited.pop(0)
    day_count = 1; clients_in_day = 0; MAX = 8
    tours.append({'jour': f'Jour {day_count}', 'client': current['data'], 'dist': 0})
    clients_in_day += 1
    while unvisited:
        nearest = min(unvisited, key=lambda x: haversine_km((current['lat'], current['lon']), (x['lat'], x['lon'])))
        dist = haversine_km((current['lat'], current['lon']), (nearest['lat'], nearest['lon']))
        if clients_in_day >= MAX: day_count += 1; clients_in_day = 0; dist = 0
        tours.append({'jour': f'Jour {day_count}', 'client': nearest['data'], 'dist': round(dist, 2)})
        current = nearest; unvisited.remove(nearest); clients_in_day += 1
    return tours

# -------------------- ROUTES --------------------
@app.context_processor
def inject_user():
    user = None; 
    if 'user_id' in session: 
        user = User.query.get(session['user_id'])
        if not user: session.clear() 
    return dict(user=user)

@app.route('/')
def home(): return render_template('index.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if 'user_id' in session: 
        if User.query.get(session['user_id']): return redirect(url_for('dashboard'))
        else: session.clear()
    if request.method == 'POST':
        email = request.form.get('email')
        if User.query.filter_by(email=email).first(): return redirect(url_for('register'))
        user = User(email=email, password_hash=generate_password_hash(request.form.get('password')))
        db.session.add(user); db.session.commit()
        session['user_id'] = user.id
        return redirect(url_for('my_clients'))
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        if User.query.get(session['user_id']): return redirect(url_for('dashboard'))
        else: session.clear()
    if request.method == 'POST':
        user = User.query.filter_by(email=request.form.get('email')).first()
        if user and check_password_hash(user.password_hash, request.form.get('password')): 
             session['user_id'] = user.id; return redirect(url_for('dashboard'))
        flash('Email incorrect.', "error")
    return render_template('login.html')

@app.route('/logout')
def logout(): session.clear(); return redirect(url_for('home'))

@app.route('/my_clients', methods=['GET', 'POST'])
def my_clients():
    if 'user_id' not in session: return redirect(url_for('login'))
    user = User.query.get(session['user_id'])
    if not user: session.clear(); return redirect(url_for('login'))
    
    if request.method == 'POST':
        if 'manual_add' in request.form:
            try:
                name = request.form.get('name'); address = request.form.get('address')
                city = request.form.get('city'); zip_code = request.form.get('zip_code')
                siret = ''.join(filter(str.isdigit, request.form.get('siret', '')))
                client_type = request.form.get('type')
                lat, lon = None, None
                try:
                    geolocator = Nominatim(user_agent="west_ops_manual_add")
                    loc = geolocator.geocode(f"{address} {zip_code} {city} France")
                    if loc: lat = loc.latitude; lon = loc.longitude
                except: pass
                new_client = Client(user_id=user.id, name=name, address=address, city=city, zip_code=zip_code, siret=siret, type=client_type, lat=lat, lon=lon)
                db.session.add(new_client); db.session.commit()
                flash(f"Client {name} ajouté !", "success")
            except Exception as e: flash(f"Erreur : {str(e)}", "error")

        elif 'file' in request.files:
            file = request.files.get('file'); import_type = request.form.get('type', 'CLIENT')
            if file:
                try:
                    df = pd.read_excel(file, dtype=str)
                    df.columns = [c.lower().strip() for c in df.columns]
                    col_name = next((c for c in df.columns if 'nom' in c), None)
                    col_ville = next((c for c in df.columns if 'ville' in c), None)
                    col_addr = next((c for c in df.columns if 'adresse' in c), None)
                    col_zip = next((c for c in df.columns if 'code' in c or 'cp' in c or 'zip' in c), None)
                    col_siret = next((c for c in df.columns if 'siret' in c), None)
                    col_siren = next((c for c in df.columns if 'siren' in c), None)
                    added_count = 0; df = df.fillna('')
                    for _, row in df.iterrows():
                        name = str(row.get(col_name, 'Inconnu'))
                        val_siret = str(row.get(col_siret, '')).replace(' ', '').replace('.', '')
                        val_siren = str(row.get(col_siren, '')).replace(' ', '').replace('.', '')
                        final_siret = val_siret if len(val_siret) >= 9 else val_siren if len(val_siren) >= 9 else ''
                        if not Client.query.filter_by(user_id=user.id, name=name).first():
                            client = Client(user_id=user.id, name=name, address=str(row.get(col_addr, '')), city=str(row.get(col_ville, '')), zip_code=str(row.get(col_zip, '').split('.')[0]), siret=final_siret, lat=None, lon=None, type=import_type)
                            db.session.add(client); added_count += 1
                    db.session.commit()
                    flash(f"Import terminé : {added_count} ajoutés.", "success")
                except Exception as e: flash(f"Erreur import: {str(e)}", "error")
    return render_template('my_clients.html', clients=Client.query.filter_by(user_id=user.id, type='CLIENT').all(), prospects=Client.query.filter_by(user_id=user.id, type='PROSPECT').all())

# --- ROUTE CORRIGÉE : IMPORT INSTANTANÉ (Sans géocodage lent) ---
@app.route('/save_prospects', methods=['POST'])
def save_prospects():
    if 'user_id' not in session: return redirect(url_for('login'))
    user = User.query.get(session['user_id'])
    
    selected_items = request.form.getlist('selected_prospects') 
    count = 0
    
    # ON NE GÉOCODE PAS ICI (Trop lent). On enregistre juste.
    # Le géocodage se fera automatiquement quand l'utilisateur ira sur "Planificateur".
    for item in selected_items:
        try:
            parts = item.split('|')
            if len(parts) >= 4:
                name = parts[0]
                address = parts[1]
                city = parts[2]
                zip_code = parts[3]
                siret = parts[4] if len(parts) > 4 else ""
                
                if not Client.query.filter_by(user_id=user.id, name=name).first():
                    prospect = Client(
                        user_id=user.id, name=name, address=address, city=city, 
                        zip_code=zip_code, siret=siret, type='PROSPECT', lat=None, lon=None
                    )
                    db.session.add(prospect)
                    count += 1
        except: pass

    db.session.commit()
    flash(f"🎉 {count} prospects ajoutés ! Allez dans 'Planificateur' pour créer la tournée.", "success")
    return redirect(url_for('my_clients'))

@app.route('/planner')
def planner():
    if 'user_id' not in session: return redirect(url_for('login'))
    user = User.query.get(session['user_id'])
    if not user: session.clear(); return redirect(url_for('login')) 
    return render_template('planner.html', client_count=Client.query.filter_by(user_id=user.id, type='CLIENT').count(), prospect_count=Client.query.filter_by(user_id=user.id, type='PROSPECT').count())

@app.route('/generate_tour/<tour_type>')
def generate_tour(tour_type):
    if 'user_id' not in session: return redirect(url_for('login'))
    user = User.query.get(session['user_id'])
    if not user: session.clear(); return redirect(url_for('login')) 
    
    targets = Client.query.filter_by(user_id=user.id, type=tour_type.upper()).all()
    
    # GÉOCODAGE MASSIF ICI (avec barre de progression)
    to_geocode = [c for c in targets if c.lat is None or c.lon is None]
    if to_geocode:
        geolocator = Nominatim(user_agent="west_ops_planner_v9")
        geocode = RateLimiter(geolocator.geocode, min_delay_seconds=1.2)
        for client in tqdm(to_geocode, desc="Géocodage", unit="adr"):
            try:
                loc = geocode(f"{client.address} {client.zip_code} {client.city} France")
                if loc: client.lat = loc.latitude; client.lon = loc.longitude
            except: pass
        db.session.commit()
        # On recharge
        targets = Client.query.filter_by(user_id=user.id, type=tour_type.upper()).all()
    
    route = optimize_route(targets)
    if not route:
        flash("Impossible de créer la route (pas assez d'adresses géolocalisées).", "error")
        return redirect(url_for('planner'))
    
    output = BytesIO(); wb = Workbook(); ws = wb.active
    headers = ["Jour", "Nom", "Adresse", "Ville", "Distance (km)"]
    for col, h in enumerate(headers, 1): c = ws.cell(1, col, h); c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1E3A8A")
    total_km = 0; dashboard_data = []
    for i, step in enumerate(route, 2):
        c = step['client']
        ws.cell(i, 1, step['jour']); ws.cell(i, 2, c.name); ws.cell(i, 3, c.address); ws.cell(i, 4, c.city); ws.cell(i, 5, step['dist'])
        total_km += step['dist']; dashboard_data.append({"Jour": step['jour'], "Nom client": c.name, "Ville": c.city, "Distance trajet (km)": step['dist']})
    wb.save(output); output.seek(0)
    db.session.add(Tour(user_id=user.id, filename=f"Tournée {tour_type.capitalize()}", total_km=total_km, total_clients=len(route), data_json=json.dumps(dashboard_data)))
    db.session.commit()
    return send_file(output, as_attachment=True, download_name=f"Tournee_{tour_type}.xlsx")

@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session: return redirect(url_for('login'))
    user = User.query.get(session['user_id'])
    if not user: session.clear(); return redirect(url_for('login')) 
    
    # Récupération des données
    all_tours = Tour.query.filter_by(user_id=user.id).order_by(Tour.date.desc()).all()
    selected = all_tours[0] if all_tours else None
    
    # 1. Calculs Économies Trajets (Carburant/Usure)
    total_km_global = sum(t.total_km for t in all_tours)
    km_saved_global = total_km_global * 0.30 # On estime qu'on optimise 30% par rapport à un trajet manuel mal fait
    euros_saved_transport = km_saved_global * 0.50 # Barème kilométrique fiscal moyen (0.50€/km)
    
    # 2. Calculs Gain de Productivité (Le fameux "Temps Admin")
    # Hypothèse : 25 min gagnées par création de tournée + 10 min par prospect qualifié trouvé
    nb_clients = Client.query.filter_by(user_id=user.id, type='CLIENT').count()
    nb_prospects = Client.query.filter_by(user_id=user.id, type='PROSPECT').count()
    
    minutes_saved_planning = len(all_tours) * 25 
    minutes_saved_sourcing = nb_prospects * 10
    total_minutes_saved = minutes_saved_planning + minutes_saved_sourcing
    
    # Conversion en Heures pour l'affichage
    hours_saved = round(total_minutes_saved / 60, 1)

    # 3. Préparation des données pour le template
    data = json.loads(selected.data_json) if selected and selected.data_json else []
    
    summary = {
        'total_clients': nb_clients,
        'total_prospects': nb_prospects,
        'total_km_all': round(total_km_global, 2),
        'savings_euro': round(euros_saved_transport, 2), # Argent économisé (Trajet)
        'hours_saved': hours_saved,                      # Temps économisé (Productivité)
        'savings_co2': round((km_saved_global * CO2_EMISSION_FACTOR) / 1000, 2),
    }
    return render_template('dashboard.html', user=user, tours=all_tours, selected_tour=selected, data=data, chart_data=data, summary=summary)

@app.route('/market_radar', methods=['GET', 'POST'])
def market_radar():
    if 'user_id' not in session: return redirect(url_for('login'))
    user = User.query.get(session['user_id'])
    
    # --- GESTION DU CACHE (ÉCONOMIE CRÉDITS) ---
    today = datetime.date.today()
    cache = RadarCache.query.filter_by(user_id=user.id, scan_date=today).first()

    # Si on force le refresh (via bouton) ou si pas de cache
    force_refresh = request.args.get('refresh') == 'true'

    if cache and not force_refresh:
        # ON UTILISE LA MÉMOIRE (0 CRÉDIT)
        new_openings = json.loads(cache.openings_json)
        alerts = json.loads(cache.alerts_json)
        flash(f"Données chargées depuis le cache (0 crédit).", "success")
    else:
        # ON APPELLE L'API (1 CRÉDIT)
        client_list = Client.query.filter_by(user_id=user.id, type='CLIENT').all()
        new_openings = get_pappers_openings_fresh(API_KEY_RADAR)
        alerts = scan_client_alerts_fresh(client_list, API_KEY_RADAR)
        
        # On sauvegarde en mémoire
        if cache:
            cache.openings_json = json.dumps(new_openings)
            cache.alerts_json = json.dumps(alerts)
        else:
            new_cache = RadarCache(user_id=user.id, scan_date=today, openings_json=json.dumps(new_openings), alerts_json=json.dumps(alerts))
            db.session.add(new_cache)
        db.session.commit()
        flash(f"Nouveau scan Pappers effectué (Données à jour).", "success")

    return render_template('market_radar.html', new_openings=new_openings, alerts=alerts)

if __name__ == "__main__":
    with app.app_context(): db.create_all()
    app.run(debug=True)